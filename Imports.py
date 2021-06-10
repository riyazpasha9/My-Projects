import csv
import datetime
import openpyxl
import os
import pythoncom
import shutil
import sys
import time
import re
import zipfile
import zlib
import glob
import win32com.client as win32
import win32com.client

sys.path.append("C:\\Users\\pashar1\\Desktop\\Test folder\\New Anaplan_API Scripts")

import no_proxyauth_Anaplan_Connector as connector

OUTPUT_PATH = "N:\\Anaplan" #path to the output directory, where generated CSVs will be stored
YARDI_FILE_PATH = "N:\\Yardi"
LOG_PATH = "N:\\Anaplan\\AIRE_YARDI.log"
#MAIL_TEMPLATE_PATH = "C:\\Users\\pashar1\\Desktop\\Real Estate - YARDI CSV Split\\Internal.msg"
today = datetime.datetime.now().date().isoformat().replace("-", "_") # in YYYY_MM_DD

AIRE_DH_MODEL = {"WorkspaceID":"" , "ModelID":""} #AIRE Data Hub model IDs
AIRE_MAIN_MODEL = {"WorkspaceID":"" , "ModelID":""} #AIRE final model IDs


OK = "SUCCESSFUL"
NOT_OK = "FAILED"

mail_body = ""

def get_timestamp():
    """Simple function to always get timestamp in the same format"""
    
    return datetime.datetime.now().isoformat()[:19] # in YYYY-MM-DDTHH:MM:SS

def log_text(text):
    """Helper to write line of text to the log"""
    
    with open(LOG_PATH, 'a') as log:
        log.write(f"[{get_timestamp()}] {text}\n")

def report(text):
    """Helper to report information to relevant streams"""

    log_text(text)
    print(text)
    
    global mail_body
    mail_body += text.replace(OK,
                              f"<span style=\"color:#00FF00;\">{OK}</span>").replace(NOT_OK,
                                                                                          f"<span style=\"color:#FF0000;\">{NOT_OK}</span>") + "<br>"

def get_latest_file():
    if not os.listdir(YARDI_FILE_PATH):
        print(f"YARDI Excel file does not exist in the given path: {YARDI_FILE_PATH}")
        send_mail()
        os.exit(-1)
    latest_file = max(glob.iglob(os.path.join(YARDI_FILE_PATH,'Anaplan_*.xlsx')), key=os.path.getctime)
    _,filename = os.path.split(latest_file)
    report(f"Input YARDI file that is being used for extracting each tab to CSV file : {filename}")
    return latest_file   

def split_workbook(LATEST_FILE):
    """Open workbook and generate CSVs from sheets"""

    if not os.path.isfile(LATEST_FILE):
        report(f"YARDI Excel file does not exist in the given path: {YARDI_FILE_PATH}")
        send_mail()
        os._exit(-1)

    YARDI_workbook = openpyxl.load_workbook(LATEST_FILE,
                                            read_only=True,
                                            data_only=True)

 
    YARDI_created = datetime.datetime.fromtimestamp(os.path.getctime(LATEST_FILE))
    YARDI_last_modified = datetime.datetime.fromtimestamp(os.path.getmtime(LATEST_FILE))
    
    report(f"YARDI Excel file successfully found and opened")
    report(f"\tYARDI created: {YARDI_created}")
    report(f"\tLast modified: {YARDI_last_modified}")

    for file in FILES_INFO:
        file_path = f"{OUTPUT_PATH}\\{file['name']}.csv"
        #try to find sheet with the right name
        try:
            worksheet = YARDI_workbook.get_sheet_by_name(file["name"])
        except KeyError:
            if not os.path.isfile(file_path):
                report(f"Action to find sheet {file['name']} in YARDI file {NOT_OK} ")
                report(f"PROCESS {NOT_OK}")
                send_mail()
                os._exit(-2)
            report("Sheet {file['name']} not found in YARDI file, but previous version of this extract is present")
            continue

        #creat CSV and copy the content of the sheet
        with open(file_path, 'w', encoding='utf-8', newline='') as output_file:
            wr = csv.writer(output_file, quoting=csv.QUOTE_ALL)
            for row in worksheet.iter_rows():
                cells = []
                for cell in row:
                    #if date, format it correctly, otherwise just save the content of the cell
                    if isinstance(cell.value, datetime.datetime):
                        cells.append(cell.value.strftime("%d/%m/%Y"))
                    else:
                        cells.append(cell.value)
                wr.writerow(cells)

        report(f"Extraction of the sheet {file['name']} {OK}, {worksheet.max_row - 1} data rows saved")


def find_file_paths():
    """Lists all files with matching names in ARCHIVE_PATH"""
    
    file_paths = []
    filenames = os.listdir(OUTPUT_PATH)
    for filename in filenames:
        if re.search(".(csv|log)$", filename):
            file_paths.append(os.path.join(OUTPUT_PATH, filename))

    return file_paths

def create_folder_copy_files(file_paths):
    """create the folderand copy the files to the directory that is created"""

    global path
    
    path = '{}\\AIRE_YARDI_SPLIT_WEEKLY_{}'.format(OUTPUT_PATH, today)
        
    if not os.path.exists(path):
        os.makedirs(path,mode=0o666)
    
         
    for file in file_paths:
        shutil.copy2(file,path)

def remove_files(file_paths):
    
    #delete files
    for file in file_paths:
        os.remove(file)

def upload_files(FILES_INFO, user_token):
    """Uploads files to the DH model"""

    for file in FILES_INFO:
        
            request = connector.upload_file(AIRE_DH_MODEL["WorkspaceID"],
                                            AIRE_DH_MODEL["ModelID"],
                                            user_token,
                                            file["id"],
                                            f"{OUTPUT_PATH}\\{file['name']}.csv")
            
            report(f"Upload {file['name']} {OK if request.ok else NOT_OK}, status code: {request.status_code}")
            
def run_actions(ACTIONS_INFO, user_token):
    """Runs actions in Anaplan models"""
    
    for action in ACTIONS_INFO:
        
            request = connector.run_anaplan_action(ACTIONS_INFO[action]["model"]["WorkspaceID"],
                                                   ACTIONS_INFO[action]["model"]["ModelID"],
                                                   user_token,
                                                   ACTIONS_INFO[action]["id"])

            report(f"Process {action} {OK if request.ok else NOT_OK}, status code: {request.status_code}")

            
def send_mail():
    pythoncom.CoInitialize()
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = "james@.com;steven@.com;kyle@.com"
    mail.CC = "riyaz.pasha@aviva.com"
    mail.Subject = f"AIRE YARDI to Anaplan Weekly upload : {get_timestamp()}"
    mail.HTMLBody = "Hello there,<br><br>"
    mail.HTMLBody += "Summary of execution:<br><br>"
    mail.HTMLBody += mail_body
    mail.HTMLBody += "<br>Regards,<br>Real Estate - YARDI Scheduled Uploads"
    mail.Send()
    log_text("Mail sent")

def main():

    log_text("===== START =====")

    LATEST_FILE = get_latest_file()

    split_workbook(LATEST_FILE) 

    user_token = connector.login_token()

    upload_files(FILES_INFO, user_token)
    run_actions(ACTIONS_INFO, user_token)

    connector.logout_token(user_token)

    send_mail()
    log_text("=====  END  =====")

    create_folder_copy_files(find_file_paths())
    remove_files(find_file_paths())
    

if __name__ == "__main__":
    

     
    #Weekly Import Names in the Model.
    #Comprehensive Tenancy Schedule-LOAD DH3_Demises from .csv,Settled Rent Reviews - LOAD DH8_RR Settlements from .csv
    #Void Schedule - LOAD DH9_Lease to from .csv, Owner Liability Schedule -LOAD DH4_Liability from .csv
     

    FILES_INFO = [
        {'name':'Comprehensive', 'id':''},
        {'name':'Reviews', 'id':''},
        {'name':'Schedule', 'id':''},
        {'name':'Owner', 'id':''}
        ]

    ACTIONS_INFO = {'1. Load Files':{'model':AIRE_DH_MODEL,'id':''},
                    'Data Load':{'model':AIRE_MAIN_MODEL,
                          'id':''}
                    }
        
    main()
        
