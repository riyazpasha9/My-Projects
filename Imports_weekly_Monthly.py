#!/usr/bin/python
# -*- coding: utf-8 -*-
"""

-------------------
Paths, which could vary for different configurations:
sys.path.append()
OUTPUT_PATH
YARDI_FILE_PATH
LOG_PATH
MAIL_TEMPLATE_PATH
"""
import argparse
import csv
import datetime
import enum
import openpyxl
import os
import pythoncom
import shutil
import sys
import time
import win32com.client

sys.path.append("C:\Anaplan_API\API_Files")

import AnaplanAPIConnector_v2_0 as connector

OUTPUT_PATH = "C:\Anaplan_API\Real Estate - YARDI Scheduled Uploads\YARDI Files Location" #path to the output directory, where generated CSVs will be stored
YARDI_FILE_PATH = "C:\Anaplan_API\Real Estate - YARDI Scheduled Uploads\YARDI Files Location\Anaplan.xlsx" #path to YARDI file
LOG_PATH = "C:\Anaplan_API\Real Estate - YARDI Scheduled Uploads\AIRE_YARDI.log" #path to the log file
MAIL_TEMPLATE_PATH = "C:\Anaplan_API\Real Estate - YARDI Scheduled Uploads\Internal.msg" #path to the template Outlook email 
AIRE_DH_MODEL = {"WorkspaceID":"" , "ModelID":""} #AIRE Data Hub model IDs
AIRE_MAIN_MODEL = {"WorkspaceID":"" , "ModelID":""} #AIRE final model IDs

OK = "SUCCESSFUL"
NOT_OK = "FAILED"

mail_body = ""

class Type(enum.IntFlag):
    WEEKLY = enum.auto()
    MONTHLY = enum.auto()
    ALL = WEEKLY | MONTHLY

def get_args():
    """Parse arguments and return type of this execution of the script"""

    #default run type
    ret_run_type = Type.ALL
    
    parser = argparse.ArgumentParser(description = "Anaplan API script to extract & upload YARDI file.")
    #get the type, in "help" print available options and default
    parser.add_argument("-t", "--run_type", help=f"provide run type (default is {ret_run_type.name}), one of: {[name for name, memeber in Type.__members__.items()]}")
    args = parser.parse_args()
    
    #if get type provided, check if it exists and return it
    if args.run_type:
        try:
            ret_run_type = Type[args.run_type.upper()]
        except KeyError:
            report('Wrong run type provided')

    report(f'Run type: {ret_run_type.name}')
    return ret_run_type

def get_timestamp():
    """Simple function to always get timestamp in the same format"""
    
    return datetime.datetime.now().isoformat()[:19] # in YYYY-MM-DDTHH:MM:SS

def log_text(text):
    """Helper to write line of text to the log"""
    
    with open(LOG_PATH, 'a') as log:
        log.write(f"[{get_timestamp()}] {text}\n")

def report(text):
    """Helper to report information to relevant streams"""

    log_text(text)
    print(text)
    
    global mail_body
    #OK -> green, NOT_OK -> red 
    mail_body += text.replace(OK,f"<span style=\"color:#00FF00;\">{OK}</span>").replace(NOT_OK,f"<span style=\"color:#FF0000;\">{NOT_OK}</span>") + "<br>"

def split_workbook(FILES_INFO, run_type):
    """Open workbook and generate CSVs from sheets"""

    if not os.path.isfile(YARDI_FILE_PATH):
        report(f"YARDI Excel file does not exist in the given path: {YARDI_FILE_PATH}")
        send_mail()
        os.exit(-1)

    YARDI_workbook = openpyxl.load_workbook(YARDI_FILE_PATH,
                                            read_only=True,
                                            data_only=True)
    
    report(f"YARDI Excel file successfully found and opened: {YARDI_FILE_PATH}")
    report(f"\tYARDI created: {datetime.datetime.fromtimestamp(os.path.getctime(YARDI_FILE_PATH))}")
    report(f"\tLast modified: {datetime.datetime.fromtimestamp(os.path.getmtime(YARDI_FILE_PATH))}")

    for file in FILES_INFO:
        if (run_type & file['type']):
            file_path = f"{OUTPUT_PATH}\\{file['name']}.csv"
            #try to find sheet with the right name
            try:
                worksheet = YARDI_workbook.get_sheet_by_name(file["name"])
            except KeyError:
                #if sheet has not been found, check if earlier file exists
                if not os.path.isfile(file_path):
                    report(f"Action to find sheet {file['name']} in YARDI file {NOT_OK}, and previous version of this extract file is missing")
                    report(f"PROCESS {NOT_OK}")
                    send_mail()
                    os.exit(-2)
                report("Sheet {file['name']} not found in YARDI, but earlier version is present (Last modified: datetime.datetime.fromtimestamp(os.path.getmtime(file_path)))")
                continue

            #creat CSV and copy the content of the sheet
            with open(file_path, 'w', encoding='utf-8', newline='') as output_file:
                wr = csv.writer(output_file, quoting=csv.QUOTE_ALL)
                for row in worksheet.iter_rows():
                    cells = []
                    for cell in row:
                        #if date, format it correctly, otherwise just save the content of the cell
                        if isinstance(cell.value, datetime.datetime):
                            cells.append(cell.value.strftime("%d/%m/%Y"))
                        else:
                            cells.append(cell.value)
                    wr.writerow(cells)

            report(f"Extraction of the sheet {file['name']} {OK}, {worksheet.max_row - 1} data rows saved")

def upload_files(FILES_INFO, user_token, run_type):
    """Uploads files to the DH model"""

    for file in FILES_INFO:
        if (run_type & file['type']):
            request = connector.upload_file(AIRE_DH_MODEL["WorkspaceID"],
                                            AIRE_DH_MODEL["ModelID"],
                                            user_token,
                                            file["id"],
                                            f"{OUTPUT_PATH}\\{file['name']}.csv")
            
            report(f"Upload {file['name']} {OK if request.ok else NOT_OK}, status code: {request.status_code}")
            
def run_actions(ACTIONS_INFO, user_token, run_type):
    """Runs actions in Anaplan models"""
    
    for action in ACTIONS_INFO:
        if (run_type & ACTIONS_INFO[action]['type']):
            request = connector.run_anaplan_action(ACTIONS_INFO[action]["model"]["WorkspaceID"],
                                                   ACTIONS_INFO[action]["model"]["ModelID"],
                                                   user_token,
                                                   ACTIONS_INFO[action]["id"])

            report(f"Process {action} {OK if request.ok else NOT_OK}, status code: {request.status_code}")

def send_mail():
    pythoncom.CoInitialize()
    Outlook = win32com.client.Dispatch("outlook.application")
    mail = Outlook.Session.OpenSharedItem(MAIL_TEMPLATE_PATH)
    mail.To = ""
    mail.Subject = f"AIRE YARDI to Anaplan upload: {get_timestamp()}"
    mail.HTMLbody = "Hello there,<br><br>"
    mail.HTMLbody += "Summary of execution:<br><br>"
    mail.HTMLbody += mail_body
    mail.HTMLbody += "<br>Regards,<br>Real Estate - YARDI Scheduled Uploads"
    mail.Send()
    log_text("Mail sent")

def main():

    log_text("===== START =====")

    run_type = get_args()

    split_workbook(FILES_INFO, run_type)
    
    user_token = connector.auto_login_token()

    upload_files(FILES_INFO, user_token, run_type)
    run_actions(ACTIONS_INFO, user_token, run_type)

    connector.logout_token(connector.make_session({"Authorization": user_token},
                                                  connector.get_proxies()))

    send_mail()

    log_text("=====  END  =====")

if __name__ == "__main__":

    FILES_INFO = [
        {'name':'MasterDirectory', 'id':'113000000054', 'type':Type.MONTHLY},
        {'name':'Property Attributes', 'id':'113000000055', 'type':Type.MONTHLY},
        {'name':'Journal Extract', 'id':'113000000027', 'type':Type.MONTHLY},
        {'name':'PayableExtract', 'id':'113000000028', 'type':Type.MONTHLY},
        {'name':'Valuation Data Extract', 'id':'113000000037', 'type':Type.MONTHLY},
        {'name':'Comprehensive Tenancy Schedule', 'id':'113000000002', 'type':Type.WEEKLY},
        {'name':'Settled Rent Reviews', 'id':'113000000022', 'type':Type.WEEKLY},
        {'name':'Void Schedule', 'id':'113000000056', 'type':Type.WEEKLY},
        {'name':'Owner Liability Schedule', 'id':'113000000003', 'type':Type.WEEKLY}
        ]
    ACTIONS_INFO = {
        '1. Load Files (Weekly)':{'model':AIRE_DH_MODEL,
                                  'id':'118000000001',
                                  'type':Type.WEEKLY},
        '1. Load Files (Monthly)':{'model':AIRE_DH_MODEL,
                                   'id':'118000000009',
                                  'type':Type.MONTHLY},
        'Full Data Load':{'model':AIRE_MAIN_MODEL,
                          'id':'118000000030',
                          'type':Type.ALL}
        }
    
    main()
